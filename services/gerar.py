from database.db import get_conn
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


OPERACAO_ENTRADA = "COMPRAS DE MERCADORIAS SEM FINANCEIRO"


def to_float(valor):
    try:
        if valor is None:
            return 0.0
        if isinstance(valor, str):
            v = valor.strip().replace(".", "").replace(",", ".")
            if v == "":
                return 0.0
            return float(v)
        return float(valor)
    except:
        return 0.0


def norm_codigo(x):
    if x is None:
        return ""
    s = str(x).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.lstrip("0")


def limpar_ncm(ncm):
    if pd.isnull(ncm):
        return ""
    return "".join(filter(str.isdigit, str(ncm)))


def buscar_aliquota_por_ncm(ncm, df_ncm):
    ncm = limpar_ncm(ncm)

    for i in range(len(ncm), 1, -1):
        prefixo = ncm[:i]
        res = df_ncm[df_ncm["ncm_limpo"] == prefixo]

        if not res.empty:
            return f"{float(res.iloc[0]['aliquota']):.2f}%"

    return "NCM não cadastrado"


def aliquota_eh_zero(valor):
    try:
        if valor is None:
            return False
        v = str(valor).replace("%", "").replace(",", ".").strip()
        if v == "":
            return False
        return float(v) == 0.0
    except:
        return False


def gerar_excel_notas(empresa_codigo, caminho_saida):

    with get_conn() as conn:

        # ✅ dados da empresa
        empresa_info = conn.execute(
            "SELECT codigo, nome FROM empresa WHERE codigo = ?",
            (empresa_codigo,)
        ).fetchone()

        estoque_df = pd.read_sql_query(
            '''
            SELECT produto, descricao, saldo_atual
            FROM "estoque"
            WHERE empresa_codigo = ?
            AND saldo_atual > 0
            ''',
            conn,
            params=(empresa_codigo,)
        )

        notas_df = pd.read_sql_query(
            '''
            SELECT *
            FROM "notas"
            WHERE empresa_codigo = ?
              AND operacao = ?
            ''',
            conn,
            params=(empresa_codigo, OPERACAO_ENTRADA)
        )

        # ✅ AGORA FILTRANDO POR EMPRESA
        ncm_rel_df = pd.read_sql_query(
            '''
            SELECT codigo_do_item, ncm, cest
            FROM "NCM E CEST"
            WHERE empresa_codigo = ?
            ''',
            conn,
            params=(empresa_codigo,)
        )

        ncm_tab_df = pd.read_sql_query(
            '''
            SELECT ncm, aliquota
            FROM ncm
            ''',
            conn
        )

    if "posicao_na_nf" in notas_df.columns:
        notas_df["posicao_na_nf"] = pd.to_numeric(
            notas_df["posicao_na_nf"], errors="coerce"
        ).fillna(0).astype(int)
    else:
        notas_df["posicao_na_nf"] = 0

    estoque_df["produto_norm"] = estoque_df["produto"].apply(norm_codigo)
    notas_df["produto_norm"] = notas_df["codigo_do_produto"].apply(norm_codigo)
    ncm_rel_df["produto_norm"] = ncm_rel_df["codigo_do_item"].apply(norm_codigo)

    ncm_rel_df["ncm_limpo"] = ncm_rel_df["ncm"].apply(limpar_ncm)
    ncm_tab_df["ncm_limpo"] = ncm_tab_df["ncm"].apply(limpar_ncm)

    grupos_notas = dict(tuple(notas_df.groupby("produto_norm")))

    resultado = []

    for _, est in estoque_df.iterrows():
        produto = est["produto"]
        descricao_estoque = est.get("descricao", "")
        produto_norm = est["produto_norm"]
        saldo = to_float(est["saldo_atual"])

        notas_prod = grupos_notas.get(produto_norm)

        if notas_prod is None or notas_prod.empty:
            resultado.append({
                "produto": produto,
                "descricao": descricao_estoque,
                "emitente": "",
                "numero": "Nota não encontrada",
                "serie": "",
                "subserie": "",
                "data_emissao": "",
                "data_entrada": "",
                "chave_acesso": "",
                "posicao_na_nf": 0,
                "quantidade_nota": 0,
                "quantidade_utilizada": 0,
                "acumulado": 0,
                "estoque": saldo,
                "cobriu_estoque": False,
                "unidade_medida": ""
            })
            continue

        notas_prod = notas_prod.sort_values(
            by=["data_de_entrada", "numero_do_documento", "posicao_na_nf"],
            ascending=[False, False, True]
        )

        acumulado = 0.0
        encontrou_nota_valida = False

        for _, nota in notas_prod.iterrows():

            numero_nota = (
                nota.get("numero_do_documento")
                or nota.get("numero")
                or nota.get("num_doc")
            )

            if not numero_nota:
                continue

            encontrou_nota_valida = True

            if acumulado >= saldo:
                break

            qtd = to_float(
                nota.get("quantidade_de_itens")
                or nota.get("quantidade")
                or nota.get("qtd")
                or 0
            )

            if qtd <= 0:
                continue

            restante = saldo - acumulado
            qtd_utilizada = min(qtd, restante)

            acumulado += qtd_utilizada

            unidade = (
                nota.get("unidade_de_medida")
                or nota.get("unidade")
                or nota.get("ucom")
                or ""
            )

            resultado.append({
                "produto": produto,
                "descricao": nota.get("descricao_produto") or descricao_estoque,
                "emitente": nota.get("nome_do_fornecedor"),
                "numero": numero_nota,
                "serie": nota.get("serie"),
                "subserie": nota.get("subserie"),
                "data_emissao": nota.get("data_de_emissao"),
                "data_entrada": nota.get("data_de_entrada"),
                "chave_acesso": nota.get("chave_nfe"),
                "posicao_na_nf": nota.get("posicao_na_nf", 0),
                "quantidade_nota": qtd,
                "quantidade_utilizada": qtd_utilizada,
                "acumulado": acumulado,
                "estoque": saldo,
                "cobriu_estoque": acumulado >= saldo,
                "unidade_medida": unidade
            })

        if not encontrou_nota_valida:
            resultado.append({
                "produto": produto,
                "descricao": descricao_estoque,
                "emitente": "",
                "numero": "Nota não encontrada",
                "serie": "",
                "subserie": "",
                "data_emissao": "",
                "data_entrada": "",
                "chave_acesso": "",
                "posicao_na_nf": 0,
                "quantidade_nota": 0,
                "quantidade_utilizada": 0,
                "acumulado": 0,
                "estoque": saldo,
                "cobriu_estoque": False,
                "unidade_medida": ""
            })

    df_final = pd.DataFrame(resultado)

    if df_final.empty:
        df_final.to_excel(caminho_saida, index=False, engine="openpyxl")
        return caminho_saida

    df_final["produto_norm"] = df_final["produto"].apply(norm_codigo)

    df_final = df_final.merge(
        ncm_rel_df[["produto_norm", "ncm", "cest", "ncm_limpo"]],
        on="produto_norm",
        how="left"
    )

    df_final["Cod. Mercadoria estoque"] = df_final["produto"]
  
    df_final["Alíquota"] = df_final["ncm_limpo"].apply(
        lambda x: buscar_aliquota_por_ncm(x, ncm_tab_df)
    )

    df_final = df_final.sort_values(
        by=["produto", "data_entrada", "posicao_na_nf"],
        ascending=[True, False, True]
    )

    df_final.to_excel(caminho_saida, index=False, engine="openpyxl")

    # ✅ formatação + cabeçalho empresa
    wb = load_workbook(caminho_saida)
    ws = wb.active

    ws.insert_rows(1)

    if empresa_info:
        ws["A1"] = f"Empresa: {empresa_info[0]} - {empresa_info[1]}"
    else:
        ws["A1"] = f"Empresa: {empresa_codigo}"

    ws["A1"].font = Font(bold=True)

    fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_amarelo = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    fonte_vermelha = Font(color="FF0000")

    headers = [cell.value for cell in ws[2]]
    idx_flag = headers.index("cobriu_estoque") + 1
    idx_aliq = headers.index("Alíquota") + 1

    for row in ws.iter_rows(min_row=3):

        if row[idx_flag - 1].value:
            for cell in row:
                cell.fill = fill_verde

        valor_aliq = row[idx_aliq - 1].value

        if str(valor_aliq).strip() == "NCM não cadastrado":
            row[idx_aliq - 1].font = fonte_vermelha
        elif aliquota_eh_zero(valor_aliq):
            row[idx_aliq - 1].fill = fill_amarelo

    wb.save(caminho_saida)

    return caminho_saida