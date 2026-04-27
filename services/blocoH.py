import pandas as pd
from openpyxl import Workbook
from database.db import get_conn


OPERACAO_ENTRADA = "COMPRAS DE MERCADORIAS SEM FINANCEIRO"


def to_float(valor):
    try:
        if valor is None:
            return 0.0
        if isinstance(valor, (int, float)):
            return float(valor)

        v = str(valor).strip()

        if v == "":
            return 0.0

        if "," in v:
            v = v.replace(".", "").replace(",", ".")
        return float(v)

    except:
        return 0.0


def norm_codigo(x):
    if x is None:
        return ""
    s = str(x).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.lstrip("0")


def gerar_bloco_h(empresa_codigo, caminho_saida):

    with get_conn() as conn:

        estoque_df = pd.read_sql_query(
            '''
            SELECT produto, descricao, saldo_atual
            FROM estoque
            WHERE empresa_codigo = ?
              AND saldo_atual > 0
            ''',
            conn,
            params=(empresa_codigo,)
        )

        notas_df = pd.read_sql_query(
            '''
            SELECT *
            FROM notas
            WHERE empresa_codigo = ?
              AND operacao = ?
            ''',
            conn,
            params=(empresa_codigo, OPERACAO_ENTRADA)
        )

    estoque_df["produto_norm"] = estoque_df["produto"].apply(norm_codigo)
    notas_df["produto_norm"] = notas_df["codigo_do_produto"].apply(norm_codigo)

    resultado = []

    grupos_notas = dict(tuple(notas_df.groupby("produto_norm")))

    for _, est in estoque_df.iterrows():

        produto = est["produto"]
        produto_norm = est["produto_norm"]
        descricao = est["descricao"]
        saldo = to_float(est["saldo_atual"])

        notas_prod = grupos_notas.get(produto_norm)

        if notas_prod is None or notas_prod.empty:
            continue

        # ✅ mesma ordenação do gerar.py
        notas_prod = notas_prod.sort_values(
            by=["data_de_entrada", "numero_do_documento", "posicao_na_nf"],
            ascending=[False, False, True]
        )

        acumulado = 0.0
        total_valor = 0.0
        total_credito = 0.0
        unidade = ""

        for _, nota in notas_prod.iterrows():

            if acumulado >= saldo:
                break

            qtd = to_float(
                nota.get("quantidade_de_itens")
                or nota.get("quantidade")
                or nota.get("qtd")
            )

            if qtd <= 0:
                continue

            restante = saldo - acumulado
            qtd_utilizada = min(qtd, restante)

            valor_unit = to_float(
                nota.get("valor_unitario")
                or nota.get("valor_unitario_do_item")
                or nota.get("vunit")
                or nota.get("preco_unitario")
            )

            total_valor += valor_unit * qtd_utilizada

            # ✅ MESMA LÓGICA DO GERAR.PY (proporcional)
            credito_total = to_float(
                nota.get("Credito - item 10")
                or nota.get("credito_item_10")
                or 0
            )

            if qtd > 0:
                total_credito += (credito_total / qtd) * qtd_utilizada

            acumulado += qtd_utilizada

            unidade = (
                nota.get("unidade_de_medida")
                or nota.get("unidade")
                or nota.get("ucom")
                or ""
            )

        if saldo == 0:
            continue

        preco_medio = total_valor / saldo

        resultado.append({
            "DESCRICAO": descricao,
            "CODIGO": produto,
            "UND": unidade,
            "QTD": saldo,
            "VALOR_UNIT": round(preco_medio, 2),
            "TOTAL_ESTOQUE": round(total_valor, 2),
            "STA": 0,
            "STB": 0,
            "CREDITO": round(total_credito, 2)
        })

    df = pd.DataFrame(resultado)

    wb = Workbook()
    ws = wb.active

    ws["C1"] = "DESCRICAO"
    ws["I1"] = "CODIGO"
    ws["J1"] = "UND"
    ws["K1"] = "QTD"
    ws["M1"] = "VALOR UNITARIO"
    ws["Q1"] = "TOTAL ESTOQUE"
    ws["T1"] = "STA"
    ws["U1"] = "STB"
    ws["X1"] = "CREDITO"

    linha = 2

    for _, row in df.iterrows():
        ws[f"C{linha}"] = row["DESCRICAO"]
        ws[f"I{linha}"] = row["CODIGO"]
        ws[f"J{linha}"] = row["UND"]
        ws[f"K{linha}"] = row["QTD"]
        ws[f"M{linha}"] = row["VALOR_UNIT"]
        ws[f"Q{linha}"] = row["TOTAL_ESTOQUE"]
        ws[f"T{linha}"] = 0
        ws[f"U{linha}"] = 0
        ws[f"X{linha}"] = row["CREDITO"]

        linha += 1

    wb.save(caminho_saida)


    return caminho_saida